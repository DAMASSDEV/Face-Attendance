import cv2, os, numpy as np
import tkinter as tk
from PIL import ImageTk, Image
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def selesai1():
    intructions.config(text="Rekam Data Telah Selesai!")
def selesai2():
    intructions.config(text="Training Wajah Telah Selesai!")
def selesai3():
    intructions.config(text="Absensi Telah Dilakukan")
def rekamDataWajah():
    wajahDir = 'datawajah'
    cam = cv2.VideoCapture(0)
    cam.set(3, 640)
    cam.set(4, 480)
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    eyeDetector = cv2.CascadeClassifier('haarcascade_eye.xml')
    faceID = entry2.get()
    nama = entry1.get()
    nim = entry2.get()
    kelas = entry3.get()
    ambilData = 1
    while True:
        retV, frame = cam.read()
        abuabu = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = faceDetector.detectMultiScale(abuabu, 1.3, 5)
        for (x, y, w, h) in faces:
            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 255), 2)
            namaFile = str(nim) +'_'+str(nama) + '_' + str(kelas) +'_'+ str(ambilData) +'.jpg'
            cv2.imwrite(wajahDir + '/' + namaFile, frame)
            ambilData += 1
            roiabuabu = abuabu[y:y + h, x:x + w]
            roiwarna = frame[y:y + h, x:x + w]
            eyes = eyeDetector.detectMultiScale(roiabuabu)
            for (xe, ye, we, he) in eyes:
                cv2.rectangle(roiwarna, (xe, ye), (xe + we, ye + he), (0, 255, 255), 1)
        cv2.imshow('webcamku', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # jika menekan tombol q akan berhenti
            break
        elif ambilData > 30:
            break
    selesai1()
    cam.release()
    cv2.destroyAllWindows()  # untuk menghapus data yang sudah dibaca

def trainingWajah():
    wajahDir = 'datawajah'
    latihDir = 'latihwajah'

    def getImageLabel(path):
        imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
        faceSamples = []
        faceIDs = []
        for imagePath in imagePaths:
            PILimg = Image.open(imagePath).convert('L')
            imgNum = np.array(PILimg, 'uint8')
            faceID = int(os.path.split(imagePath)[-1].split('_')[0])
            faces = faceDetector.detectMultiScale(imgNum)
            for (x, y, w, h) in faces:
                faceSamples.append(imgNum[y:y + h, x:x + w])
                faceIDs.append(faceID)
        return faceSamples, faceIDs  # Pindahkan return ke sini


    faceRecognizer = cv2.face.LBPHFaceRecognizer_create()
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    faces, IDs = getImageLabel(wajahDir)
    faceRecognizer.train(faces, np.array(IDs))
    # simpan
    faceRecognizer.write(latihDir + '/training.xml')
    selesai2()



def markAttendance(name):
    # Buka atau buat file Excel
    workbook = openpyxl.load_workbook('Attendance.xlsx') if os.path.exists('Attendance.xlsx') else openpyxl.Workbook()
    sheet = workbook.active

    # Cek apakah sheet kosong, tambahkan header jika belum ada
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(['Mata Kuliah: Pengantar Sains Data'])  # Judul mata kuliah di baris 3
            sheet.append(['Tanggal Absen:'])  # Tanggal absen di baris 2
            sheet.append(['Nama', 'Kelas', 'NPM', 'Waktu Kedatangan'])

        
    # Cek apakah nama sudah ada di sheet
    namelist = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    if name not in namelist:
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')  # Waktu absensi
        yournim = entry2.get()
        yourclass = entry3.get()

        # Tambahkan data baru ke sheet
        sheet.append([name, yourclass, yournim, dtString])

        # Atur lebar kolom otomatis
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    dateList = [sheet.cell(row=i, column=5).value for i in range(6, sheet.max_row + 1)]
    dateString = now.strftime('%Y-%m-%d')  # Tanggal sekarang
    dateList = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]  # Cek dari baris 2
    if dateString not in dateList:
        sheet.insert_rows(2)  # Sisipkan baris kedua untuk tanggal
        sheet.cell(row=4, column=2, value=dateString)  # Tambahkan tanggal di baris kedua


        # Tambahkan data baru ke sheet

    # Simpan file Excel
    workbook.save('Attendance.xlsx')




def absensiWajah():
    wajahDir = 'datawajah'
    latihDir = 'latihwajah'
    cam = cv2.VideoCapture(0)
    cam.set(3, 640)
    cam.set(4, 480)
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    faceRecognizer = cv2.face.LBPHFaceRecognizer_create()
    faceRecognizer.read(latihDir + '/training.xml')
    font = cv2.FONT_HERSHEY_SIMPLEX

    #id = 0
    yourname = entry1.get()
    names = []
    names.append(yourname)
    minWidth = 0.1 * cam.get(3)
    minHeight = 0.1 * cam.get(4)

    while True:
        retV, frame = cam.read()
        frame = cv2.flip(frame, 1)
        abuabu = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = faceDetector.detectMultiScale(abuabu, 1.2, 5, minSize=(round(minWidth), round(minHeight)), )
        for (x, y, w, h) in faces:
            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0),2)
            id, confidence = faceRecognizer.predict(abuabu[y:y+h,x:x+w])
            if (confidence < 100):
                id = names[0]
                confidence = "  {0}%".format(round(150 - confidence))
            elif confidence < 50:
                id = names[0]
                confidence = "  {0}%".format(round(170 - confidence))

            elif confidence > 70:
                id = "Tidak Diketahui"
                confidence = "  {0}%".format(round(150 - confidence))

            cv2.putText(frame, str(id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
            cv2.putText(frame, str(confidence), (x + 5, y + h + 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 0), 2)

        cv2.imshow('ABSENSI WAJAH', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):  # jika menekan tombol q akan berhenti
            break
    markAttendance(id)
    selesai3()
    cam.release()
    cv2.destroyAllWindows()

# GUI
root = tk.Tk()
# mengatur canvas (window tkinter)
canvas = tk.Canvas(root, width=700, height=400)
canvas.grid(columnspan=3, rowspan=8)
canvas.configure(bg="#7AB2D3")
# judul
judul = tk.Label(root, text="Face Attendance - Smart Absensi", font=("Roboto",34),bg="#4A628A", fg="black")
canvas.create_window(350, 80, window=judul)
#credit
made = tk.Label(root, text="PERKENALKAN KAMI DARI KELOMPOK 8", font=("Times New Roman",13), bg="#7AB2D3",fg="black")
canvas.create_window(360, 20, window=made)
# for entry data nama
entry1 = tk.Entry (root, font="Roboto")
canvas.create_window(457, 170, height=25, width=411, window=entry1)
label1 = tk.Label(root, text="Nama Siswa", font="Roboto", fg="black", bg="#7AB2D3")
canvas.create_window(90,170, window=label1)
# for entry data nim
entry2 = tk.Entry (root, font="Roboto")
canvas.create_window(457, 210, height=25, width=411, window=entry2)
label2 = tk.Label(root, text="NPM", font="Roboto", fg="black", bg="#7AB2D3")
canvas.create_window(60, 210, window=label2)
# for entry data kelas
entry3 = tk.Entry (root, font="Roboto")
canvas.create_window(457, 250, height=25, width=411, window=entry3)
label3 = tk.Label(root, text="Kelas", font="Roboto", fg="black", bg="#7AB2D3")
canvas.create_window(65, 250, window=label3)

global intructions

# tombol untuk rekam data wajah
intructions = tk.Label(root, text="MELENGKAPI TUGAS PENGATAR SAINS DATA", font=("Roboto",15),fg="black",bg="#7AB2D3")
canvas.create_window(370, 300, window=intructions)
Rekam_text = tk.StringVar()
Rekam_btn = tk.Button(root, textvariable=Rekam_text, font="Roboto", bg="#20bebe", fg="white", height=1, width=15,command=rekamDataWajah)
Rekam_text.set("Take Images")
Rekam_btn.grid(column=0, row=7)

# tombol untuk training wajah
Rekam_text1 = tk.StringVar()
Rekam_btn1 = tk.Button(root, textvariable=Rekam_text1, font="Roboto", bg="#20bebe", fg="white", height=1, width=15,command=trainingWajah)
Rekam_text1.set("Training")
Rekam_btn1.grid(column=1, row=7)

# tombol absensi dengan wajah
Rekam_text2 = tk.StringVar()
Rekam_btn2 = tk.Button(root, textvariable=Rekam_text2, font="Roboto", bg="#20bebe", fg="white", height=1, width=20, command=absensiWajah)
Rekam_text2.set("Automatic Attendance")
Rekam_btn2.grid(column=2, row=7)

root.mainloop()



# FUNGSI TOMBOL

# - Take image ==> Mengambil data agar bisa di gunakan
# - Training ==> Melatih komputer agar paham atau mengenali wajah kita 